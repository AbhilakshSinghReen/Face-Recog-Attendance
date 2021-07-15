import mysql.connector
import os
import openpyxl
from datetime import datetime

dirname = os.path.dirname(__file__)
ExportWorkbook=openpyxl.load_workbook(os.path.join(dirname,"Export/Export.xlsx"))
#ExportWorkbook=openpyxl.Workbook()
#ExportWorkbook.create_sheet("Sheet1")
ExportSheet=ExportWorkbook["Sheet1"]

for i in range(1,ExportSheet.max_row+1):
    for j in range(1,ExportSheet.max_column+1):
        ExportSheet.cell(row=i,column=j).value=None



class ExportTuple:
    def __init__(self,ID,Name,Presents,Absents):
        self.ID=ID
        self.Name=Name
        self.Presents=Presents
        self.Absents=Absents

    def ChangeValues(self,NewID,NewName,NewPresents,NewAbsents):
        self.ID=NewID
        self.Name=NewName
        self.Presents=NewPresents
        self.Absents=NewAbsents

    def PrintDataForDebugging(self):
        print("ID = "+self.ID)
        print("Name = "+self.Name)
        print("Presents = "+self.Presents)
        print("Absents = "+self.Absents)
        print("")

class ExportTupleForSingleDate:
    def __init__(self,ID,Name,Status,InTime,OutTime):
        self.ID=ID
        self.Name=Name
        self.Status=Status
        self.InTime=InTime
        self.OutTime=OutTime

    def PrintDataForDebugging(self):
        print("ID = ",self.ID)
        print("Name = ",self.Name)
        print("Status = ",self.Status)
        print("InTime = ",self.InTime)
        print("OutTime = ",self.OutTime)

class ExportTupleForSingleName:
    def __init__(self,Date,ID,Name,Status,InTime,OutTime):
        self.Date=Date
        self.ID=ID
        self.Name=Name
        self.Status=Status
        self.InTime=InTime
        self.OutTime=OutTime



dirname = os.path.dirname(__file__)
os.system("python "+os.path.join(dirname,'DbManagement.py'))

fh=open(os.path.join(dirname,'Config/host.txt'),'r')
HostName=str(fh.readline())
fh.close()
fh=open(os.path.join(dirname,'Config/user.txt'),'r')
UserName=str(fh.readline())
fh.close()
fh=open(os.path.join(dirname,'Config/password.txt'),'r')
Password=str(fh.readline())
fh.close()
HostName=HostName.replace("\n","")
UserName=UserName.replace("\n","")
Password=Password.replace("\n","")

MainDB= mysql.connector.connect(host=HostName,user=UserName,passwd=Password,auth_plugin='mysql_native_password')
MainCursor=MainDB.cursor(buffered=True)

try:
    MainCursor.execute("use face__recog")
except:
    print("hllu")


fh=open(os.path.join(dirname,'Export/ExportDetails.txt'),'r')
Name=str(fh.readline())
Name=Name.replace("\n","")
Method=str(fh.readline())
Method=Method.replace("\n","")
StartDate=str(fh.readline())
EndDate=str(fh.readline())

FilePath=str(fh.readline())

TargetLevel=str(fh.readline())
TargetEmploymentStatus=str(fh.readline())
FileName=str(fh.readline())

fh.close()

FilePath=FilePath.strip()
TargetLevel=TargetLevel.strip()
TargetEmploymentStatus=TargetEmploymentStatus.strip()
FileName=FileName.strip()

IDs=[]
Data=[]
SingleDayData=[]
SingleIdData=[]

NewDatae=[]
NewSingleDayDatae=[]

def GetNameFrom_ID(ID):
    MainCursor.execute("select Full_Name from employees where id = '"+str(ID)+"'")
    for i in MainCursor:
        i=str(i)
        i=i.replace("'","")
        i=i.replace(")","")
        i=i.replace("(","")
        i=i.replace(",","")
        #print(i)
        return i

def GetLevelFrom_ID(ID):
    MainCursor.execute("select Title from employees where id = '"+str(ID)+"'")
    for i in MainCursor:
        i=str(i)
        i=i.replace("'","")
        i=i.replace(")","")
        i=i.replace("(","")
        i=i.replace(",","")
        print("returning level = ",i)


        return i

def GetEmploymentStatusFrom_ID(ID):
    MainCursor.execute("select Employment_Status from employees where id = '"+str(ID)+"'")
    for i in MainCursor:
        i=str(i)
        i=i.replace("'","")
        i=i.replace(")","")
        i=i.replace("(","")
        i=i.replace(",","")
        #print(i)
        return i



print(TargetLevel)
print(TargetEmploymentStatus)
if(Name==""):
    MainCursor.execute("select distinct id from employees")
    for i in MainCursor:
        i=str(i)
        i=i.replace(",)","")
        i=i.replace("(","")
        i=i.replace("'","")
        print(i)
        IDs.append(i)

    for i in IDs:
        NewData=ExportTuple(str(i),str(GetNameFrom_ID(i)),"0","0")
        Data.append(NewData)

    if(TargetLevel!="Any" and TargetEmploymentStatus!="Any"):
        for XData in Data:
            if (GetLevelFrom_ID(XData.ID) == TargetLevel and GetEmploymentStatusFrom_ID(XData.ID) ==TargetEmploymentStatus):
                NewDatae.append(XData)
    elif(TargetLevel!="Any" and TargetEmploymentStatus=="Any"):
        for XData in Data:
            if (GetLevelFrom_ID(XData.ID) == TargetLevel):
                NewDatae.append(XData)
    elif(TargetLevel=="Any" and TargetEmploymentStatus!="Any"):
        for XData in Data:
            if (GetEmploymentStatusFrom_ID(XData.ID)==TargetEmploymentStatus):
                NewDatae.append(XData)
    else:
        for XData in Data:
            NewDatae.append(XData)

    print("NUmber of employees=",len(Data))
else:
    MainCursor.execute("select distinct id from employees where id='"+Name+"'")
    for i in MainCursor:
        i=str(i)
        i=i.replace(",)","")
        i=i.replace("(","")
        i=i.replace("'","")
        print(i)
        IDs.append(i)

    for i in IDs:
        NewData=ExportTuple(str(i),str(GetNameFrom_ID(i)),"0","0")
        Data.append(NewData)

    print("NUmber of employees=",len(Data))


if not (StartDate==EndDate):
   if not(len(IDs)==1):
    StartDate=StartDate+" 00:00:00"
    EndDate=EndDate+" 23:59:59"
    ExportSheet.cell(row=1,column=1).value="ID"
    ExportSheet.cell(row=1,column=2).value="Name"
    ExportSheet.cell(row=1,column=3).value="Presents"
    ExportSheet.cell(row=1,column=4).value="Absents"
    RowNumber=2

    for DataX in NewDatae:
                    MainCursor.execute("select count(*) from people where status='Present' and ID='"+DataX.ID+"' and InTime>'"+StartDate+"' and OutTime<='"+EndDate+"'")
                    for i in MainCursor:
                        i=str(i)
                        i=i.replace(",)","")
                        i=i.replace("(","")
                        i=i.replace("'","")
                        DataX.ChangeValues(DataX.ID,DataX.Name,i,"0")

                    MainCursor.execute("select count(*) from people where status='Absent' and ID='"+DataX.ID+"' and InTime>'"+StartDate+"' and OutTime<='"+EndDate+"'")
                    for i in MainCursor:
                        i=str(i)
                        i=i.replace(",)","")
                        i=i.replace("(","")
                        i=i.replace("'","")
                        DataX.ChangeValues(DataX.ID,DataX.Name,DataX.Presents,i)

                    DataX.PrintDataForDebugging()
                    ExportSheet.cell(row=RowNumber,column=1).value=str(DataX.ID)
                    ExportSheet.cell(row=RowNumber,column=2).value=str(DataX.Name)
                    ExportSheet.cell(row=RowNumber,column=3).value=str(DataX.Presents)
                    ExportSheet.cell(row=RowNumber,column=4).value=str(DataX.Absents)
                    RowNumber+=1

   else:
       ExportSheet.cell(row=1,column=1).value="Date"
       ExportSheet.cell(row=1,column=2).value="ID"
       ExportSheet.cell(row=1,column=3).value="Name"
       ExportSheet.cell(row=1,column=4).value="Status"
       ExportSheet.cell(row=1,column=5).value="InTime"
       ExportSheet.cell(row=1,column=6).value="OutTime"
       MainCursor.execute("select ID,Name,Status,InTime,OutTime from people where ID='"+str(IDs[0])+"' and EntryNumber IS NOT NULL")
       for i in MainCursor:
           i=str(i)
           i=i.replace("datetime.datetime","")
           #print(i)
           i=i.replace("'","")
           i=i.replace(", ",",")
           i=i.replace("(","")
           i=i.replace(")","")
           #print(i)
           TempData=i.split(",")
           TempID=TempData[0]
           TempName=TempData[1]
           TempStatus=TempData[2]
           TempInTime="-"
           TempOutTime="-"
           TempDate=TempData[3]+"-"+TempData[4]+"-"+TempData[5]

           #print("Status = ",TempStatus)
           #print("Length = ",len(TempData))
           #print("")
           if(TempStatus=="Present"):
               if(len(TempData)==15):
                   TempInTime=TempData[3]+"-"+TempData[4]+"-"+TempData[5]+" "+TempData[6]+":"+TempData[7]+":"+TempData[8]
                   TempOutTime=TempData[9]+"-"+TempData[10]+"-"+TempData[11]+" "+TempData[12]+":"+TempData[13]+":"+TempData[14]
               elif(len(TempData[8])==4):
                   TempInTime=TempData[3]+"-"+TempData[4]+"-"+TempData[5]+" "+TempData[6]+":"+TempData[7]+":"+"00"
                   TempOutTime=TempData[8]+"-"+TempData[9]+"-"+TempData[10]+" "+TempData[11]+":"+TempData[12]+":"+TempData[13]
               else:
                   TempInTime=TempData[3]+"-"+TempData[4]+"-"+TempData[5]+" "+TempData[6]+":"+TempData[7]+":"+TempData[8]
                   TempOutTime=TempData[9]+"-"+TempData[10]+"-"+TempData[11]+" "+TempData[12]+":"+TempData[13]+":"+"00"

           SingleIdData.append(ExportTupleForSingleName(TempDate,TempID,TempName,TempStatus,TempInTime,TempOutTime))

           RowNumber=2
           for DataX in SingleIdData:
               #DataX.PrintDataForDebugging()
               ExportSheet.cell(row=RowNumber,column=1).value=str(DataX.Date)
               ExportSheet.cell(row=RowNumber,column=2).value=str(DataX.ID)
               ExportSheet.cell(row=RowNumber,column=3).value=str(DataX.Name)
               ExportSheet.cell(row=RowNumber,column=4).value=str(DataX.Status)
               ExportSheet.cell(row=RowNumber,column=5).value=str(DataX.InTime)
               ExportSheet.cell(row=RowNumber,column=6).value=str(DataX.OutTime)
               RowNumber+=1




else:
    ExportSheet.cell(row=1,column=1).value="ID"
    ExportSheet.cell(row=1,column=2).value="Name"
    ExportSheet.cell(row=1,column=3).value="Status"
    ExportSheet.cell(row=1,column=4).value="InTime"
    ExportSheet.cell(row=1,column=5).value="OutTime"

    StartDate=StartDate+" 00:00:00"
    EndDate=EndDate+" 23:59:59"
    #print("select ID,Name,Status,InTime,OutTime from people where InTime>'"+StartDate+"' and OutTime<='"+EndDate+"'")
    if(Name==""):
        MainCursor.execute("select ID,Name,Status,InTime,OutTime from people where InTime>='"+StartDate+"' and OutTime<='"+EndDate+"'")
    else:
        MainCursor.execute("select ID,Name,Status,InTime,OutTime from people where InTime>='"+StartDate+"' and OutTime<='"+EndDate+"' and ID='"+Name+"'")
    for i in MainCursor:
        i=str(i)
        i=i.replace("datetime.datetime","")
        #print(i)
        i=i.replace("'","")
        i=i.replace(", ",",")
        i=i.replace("(","")
        i=i.replace(")","")
        #print(i)
        TempData=i.split(",")
        TempID=TempData[0]
        TempName=TempData[1]
        TempStatus=TempData[2]
        TempInTime="-"
        TempOutTime="-"
        if(TempStatus=="Present"):
            TempInTime=TempData[3]+"-"+TempData[4]+"-"+TempData[5]+" "+TempData[6]+":"+TempData[7]+":"+TempData[8]
            TempOutTime=TempData[9]+"-"+TempData[10]+"-"+TempData[11]+" "+TempData[12]+":"+TempData[13]+":"+TempData[14]

        SingleDayData.append(ExportTupleForSingleDate(TempID,TempName,TempStatus,TempInTime,TempOutTime))

    if(TargetLevel!="Any" and TargetEmploymentStatus!="Any"):
        for X in SingleDayData:
            if (GetLevelFrom_ID(X.ID) == TargetLevel and GetEmploymentStatusFrom_ID(X.ID) ==TargetEmploymentStatus):
                NewSingleDayDatae.append(X)
    elif(TargetLevel!="Any" and TargetEmploymentStatus=="Any"):
        for X in SingleDayData:
            if (GetLevelFrom_ID(X.ID) == TargetLevel):
                NewSingleDayDatae.append(X)
    elif(TargetLevel=="Any" and TargetEmploymentStatus!="Any"):
        for X in SingleDayData:
            if (GetEmploymentStatusFrom_ID(X.ID)==TargetEmploymentStatus):
                NewSingleDayDatae.append(X)
    else:
        for X in SingleDayData:
            NewSingleDayDatae.append(X)


    RowNumber=2
    for DataX in NewSingleDayDatae:
        DataX.PrintDataForDebugging()

        ExportSheet.cell(row=RowNumber,column=1).value=str(DataX.ID)
        ExportSheet.cell(row=RowNumber,column=2).value=str(DataX.Name)
        ExportSheet.cell(row=RowNumber,column=3).value=str(DataX.Status)
        ExportSheet.cell(row=RowNumber,column=4).value=str(DataX.InTime)
        ExportSheet.cell(row=RowNumber,column=5).value=str(DataX.OutTime)
        RowNumber+=1

        print("")
CurrentTime=str(datetime.now())
CurrentTime=CurrentTime[:19]
CurrentTime=CurrentTime.replace(":","-")
#print(CurrentTime)

if(FileName==""):
    ExportWorkbook.save(os.path.join(FilePath,"Export ["+CurrentTime+"].xlsx"))
else:
    ExportWorkbook.save(os.path.join(FilePath,FileName+".xlsx"))
